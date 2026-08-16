"""
v2_map_genes.py - V2A MAPPING STAGE ONLY (no training, no split).

Maps each WT-SpCas9 guide from the DeepHF efficiency file (Supplementary Data 2,
'deepHF_clean.csv' or the raw xlsx) to its AUTHORITATIVE target gene using the
original DeepHF design table (Supplementary Data 1, MOESM2_ESM.xlsx).

Mapping rule (verified against the paper, PMC6753114):
  each design oligo was designed FOR one gene; the guide is the leading
  20-21 nt uppercase run of the oligo; the gene label = (Symbol, Entrez ID)
  of that design row. We join efficiency guides to design guides on the
  20-nt spacer sequence. The design oligo ALSO embeds the guide+PAM target
  site; we cross-check that embedded PAM equals the efficiency PAM column.

PAM column: sourced from Supplementary Data 2 'PAM' (joined by gRNA_Seq),
  cross-checked against the guide+PAM embedded in the design oligo.

Outputs (V2 namespace only; V1 untouched):
  results/v2/gene_labels.csv      - one row per retained guide
  results/v2/mapping_report.txt   - full mapping QC report

Exclusions (documented, not inferred):
  - unmatched guides (not found in design table)
  - ambiguous guides (20-mer maps to >1 gene)
  - malformed rows (invalid symbols / bad sequences)
No transcriptome matching is used.
"""
import argparse, collections, datetime, os, re
import numpy as np
import pandas as pd
import openpyxl

GUIDE_RE = re.compile(r'^([ACGT]{20,21})')
SYMBOL_RE = re.compile(r'^[A-Za-z][A-Za-z0-9\-\.]*$')


def valid_symbol(sym):
    """A design-table symbol is valid if it looks like a gene symbol
    (and is not an Excel auto-format date artifact)."""
    if sym is None:
        return False
    s = str(sym).strip()
    if not SYMBOL_RE.fullmatch(s):
        return False
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            datetime.datetime.strptime(s, fmt)
            return False
        except ValueError:
            pass
    return True


def load_design_map(xlsx_path):
    """Build {20mer: {(symbol, entrez_id)}} + {20mer: {embedded PAM}} from Supp Data 1.

    Gene matching is IDENTICAL to before; the embedded-PAM map is additive
    (used only for the PAM cross-check in the report)."""
    design = collections.defaultdict(set)
    embedded_pam = collections.defaultdict(set)
    malformed = collections.Counter()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        sym, gid, seq = row[0], row[1], row[2]
        if seq is None or str(seq).strip() == '':
            malformed['empty_seq'] += 1
            continue
        m = GUIDE_RE.match(str(seq))
        if not m:
            malformed['bad_seq'] += 1
            continue
        pre = m.group(1)
        g20s = {pre[:20]}
        if len(pre) == 21:
            g20s.add(pre[1:21])
        if not valid_symbol(sym):
            malformed['invalid_symbol'] += 1
            continue
        for g20 in g20s:
            design[g20].add((str(sym).strip(), str(gid).strip()))
            mm = re.search(re.escape(g20) + r'([ACGT]GG)', str(seq))
            if mm:
                embedded_pam[g20].add(mm.group(1))
    return design, dict(malformed), dict(embedded_pam)


def load_pam_lookup(xlsx_path):
    """{gRNA_Seq: PAM} from Supplementary Data 2 (the authoritative PAM source)."""
    lookup = {}
    if not xlsx_path or not os.path.exists(xlsx_path):
        return lookup
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=3, values_only=True):
        seq, pam = r[2], r[3]
        if seq is not None and pam is not None:
            lookup[str(seq).upper()] = str(pam).upper()
    return lookup


def load_efficiency(clean_csv=None, xlsx_path=None, pam_lookup=None):
    """Load efficiency rows as (gRNA_Seq, PAM, Wt_Efficiency).

    When reading from the cleaned CSV (which lacks PAM), PAM is filled from
    the Supplementary Data 2 lookup keyed by gRNA_Seq."""
    rows = []
    pam_lookup = pam_lookup or {}
    if clean_csv and os.path.exists(clean_csv):
        df = pd.read_csv(clean_csv)
        for s, e in zip(df.iloc[:, 0], df.iloc[:, 1]):
            seq = str(s).upper()
            rows.append((seq, pam_lookup.get(seq, ''), e))
    else:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(min_row=3, values_only=True):
            seq, pam, eff = r[2], r[3], r[5]
            if seq is None or eff is None:
                continue
            rows.append((str(seq).upper(), str(pam).upper(), eff))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--efficiency-xlsx',
                    default='41467_2019_12281_MOESM3_ESM.xlsx',
                    help='Supplementary Data 2 (efficiency). Used if deepHF_clean.csv absent.')
    ap.add_argument('--clean-csv', default='data/deepHF_clean.csv',
                    help='Your cleaned efficiency CSV (Sequence,Activity).')
    ap.add_argument('--design-xlsx', default='data/deephf_design.xlsx',
                    help='Supplementary Data 1 (design table, MOESM2).')
    ap.add_argument('--out', default='results/v2', help='Output dir (v2 namespace).')
    args = ap.parse_args()

    design, malformed, embedded_pam = load_design_map(args.design_xlsx)
    pam_lookup = load_pam_lookup(args.efficiency_xlsx)
    if pam_lookup:
        print(f"PAM lookup from Supp Data 2: {len(pam_lookup):,} guides")
    else:
        print("WARNING: efficiency xlsx not found - PAM will be empty")
    eff = load_efficiency(clean_csv=args.clean_csv, xlsx_path=args.efficiency_xlsx,
                          pam_lookup=pam_lookup)
    print(f"Design table: {len(design):,} unique 20-mers (malformed rows: {malformed})")
    print(f"Efficiency rows: {len(eff):,}")

    # dedupe efficiency by gRNA_Seq (mirror of V1 cleaning)
    seen, uniq = set(), []
    for s, p, e in eff:
        if s not in seen:
            seen.add(s)
            uniq.append((s, p, e))
    total = len(uniq)
    print(f"Unique WT guides: {total:,}")

    matched = 0
    unmatched, ambiguous, retained = [], [], []
    for s, p, e in uniq:
        gs = design.get(s)
        if not gs:
            unmatched.append((s, p, e))
        elif len(gs) > 1:
            ambiguous.append((s, p, e, gs))
        else:
            sym, gid = next(iter(gs))
            retained.append((s, p, e, sym, gid))

    print(f"Matched: {len(retained)+len(ambiguous):,} | unmatched: {len(unmatched):,} | "
          f"ambiguous(>1 gene): {len(ambiguous):,} | retained: {len(retained):,}")

    # genes
    gene_counts = collections.Counter((sym, gid) for _, _, _, sym, gid in retained)
    genes = len(gene_counts)

    # save
    os.makedirs(args.out, exist_ok=True)
    df_out = pd.DataFrame(retained, columns=['Sequence', 'PAM', 'Activity', 'Gene_Symbol', 'Entrez_ID'])
    df_out.to_csv(os.path.join(args.out, 'gene_labels.csv'), index=False)

    report = []
    report.append("=" * 70)
    report.append("V2A MAPPING REPORT - DeepHF efficiency <-> design table (Supp Data 1)")
    report.append("=" * 70)
    report.append("")
    report.append("MATCHING FIELDS USED")
    report.append("  efficiency: gRNA_Seq (20-nt spacer, from 'deepHF_clean.csv' or Supp Data 2)")
    report.append("  design    : leading 20-21 nt uppercase run of oligo 'Seq' (guide spacer)")
    report.append("  gene label: (Symbol, Entrez ID) of the design row the oligo was made for")
    report.append("  PAM       : from Supp Data 2 'PAM' column, joined by gRNA_Seq;")
    report.append("              cross-checked vs guide+PAM embedded in design oligo")
    report.append("")
    report.append("MAPPING RESULTS")
    report.append(f"  total unique WT guides         : {total:,}")
    n_matched = len(retained) + len(ambiguous)
    report.append(f"  matched to design table        : {n_matched:,} ({n_matched/total*100:.2f}%)")
    report.append(f"    - unambiguous (1 gene)        : {len(retained):,} ({len(retained)/total*100:.2f}%)")
    report.append(f"    - ambiguous (>1 gene)         : {len(ambiguous):,} ({len(ambiguous)/total*100:.2f}%)")
    report.append(f"  unmatched (excluded)            : {len(unmatched):,} ({len(unmatched)/total*100:.2f}%)")
    report.append(f"  malformed design rows skipped   : {malformed}")
    report.append(f"  guides retained for V2A         : {len(retained):,}")
    report.append(f"  genes represented               : {genes:,}")
    report.append(f"  guides/gene: median {np.median(list(gene_counts.values())):.0f}, "
                  f"mean {np.mean(list(gene_counts.values())):.1f}, max {max(gene_counts.values())}")
    report.append(f"  genes with >=5 guides           : {sum(1 for v in gene_counts.values() if v >= 5):,}")
    report.append("")
    # ---- PAM verification block ----
    missing_pam = [x for x in retained if not x[1]]
    cross_ok = cross_mm = cross_none = 0
    for s, p, e, sym, gid in retained:
        emb = embedded_pam.get(s)
        if not emb:
            cross_none += 1
        elif p in emb:
            cross_ok += 1
        else:
            cross_mm += 1
    report.append("PAM COLUMN VERIFICATION")
    report.append(f"  retained guides with non-empty PAM : {len(retained)-len(missing_pam):,} / {len(retained):,}")
    report.append("  PAM cross-check vs design-oligo embedded guide+PAM:")
    report.append(f"    matched            : {cross_ok:,}")
    report.append(f"    mismatch           : {cross_mm:,}")
    report.append(f"    no embedded found  : {cross_none:,}")
    report.append("")
    report.append("AMBIGUOUS GUIDES (excluded, listed)")
    for s, p, e, gs in sorted(ambiguous):
        report.append(f"  {s}  PAM={p}  genes={sorted(gs)}")
    report.append("")
    report.append("UNMATCHED SAMPLE (first 20)")
    for s, p, e in unmatched[:20]:
        report.append(f"  {s}  PAM={p}")
    report.append("")

    with open(os.path.join(args.out, 'mapping_report.txt'), 'w') as f:
        f.write('\n'.join(report))
    print(f"\nSaved {os.path.join(args.out, 'gene_labels.csv')} and "
          f"{os.path.join(args.out, 'mapping_report.txt')}")


if __name__ == '__main__':
    main()
